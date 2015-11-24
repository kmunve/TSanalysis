#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import print_function

'''
### Column names
- StasjonsID: e.g. 51346
- stasjonsNavn_kort: FILEFJELL   //Mellomrom og andre spesialkarakterer er tatt bort. Det er dette navnet Dagrunn gir til filnavnet. Husker at det var litt forskjell av og til.
- stasjonsNavn: FILEFJELL
- latitude: 61.178
- longitude: 8.1125
- UTM_EAST: 130049
- UTM_NORTH: 6802130
- URL: ftp://ftp.met.no/users/dagrunvs/xgeo/    //Skal det enda kjøres fra Dagrunns konto?
- MOH:


__author__ = 'kmu'
'''

import openpyxl
from crocus_station_db import CrocusStationDB

ftp_url = r'''ftp://ftp.met.no/users/dagrunvs/xgeo/'''

db = CrocusStationDB('../Test/Data/stations.db')
all_stations = db.get_all_stations()


wb = openpyxl.Workbook()
ws = wb.get_active_sheet()
ws.title = 'crocus_stations'

title_row = 0

stnr_col = 0
name_short_col = 1
name_col = 2
lat_col = 3
lon_col = 4
utme_col = 5
utmn_col = 6
url_col = 7
amsl_col = 8


cA = ws.cell(row=title_row, column=stnr_col)
cA.value = 'StasjonsID'

cB = ws.cell(row=title_row, column=name_short_col)
cB.value = 'stasjonsNavn_kort'

cC = ws.cell(row=title_row, column=2)
cC.value = 'stasjonsNavn'

cD = ws.cell(row=title_row, column=3)
cD.value = 'latitude'

cE = ws.cell(row=title_row, column=4)
cE.value = 'longitude'

cF = ws.cell(row=title_row, column=5)
cF.value = 'UTM_EAST'

cG = ws.cell(row=title_row, column=6)
cG.value = 'UTM_NORTH'

cH = ws.cell(row=title_row, column=7)
cH.value = 'URL'

cI = ws.cell(row=title_row, column=8)
cI.value = 'MOH'

for station in all_stations:
    c = ws.cell(row=station[0], column=stnr_col)
    c.value = station[1]

    c = ws.cell(row=station[0], column=name_short_col)
    c.value = station[2].split(' ')[0]

    c = ws.cell(row=station[0], column=name_col)
    c.value = station[2]

    c = ws.cell(row=station[0], column=lat_col)
    c.value = station[3]

    c = ws.cell(row=station[0], column=lon_col)
    c.value = station[4]

    c = ws.cell(row=station[0], column=utme_col)
    c.value = station[5]

    c = ws.cell(row=station[0], column=utmn_col)
    c.value = station[6]

    c = ws.cell(row=station[0], column=url_col)
    c.value = ftp_url

    c = ws.cell(row=station[0], column=amsl_col)
    c.value = station[7]

wb.save('../Test/Data/crocus_stations.xlsx')


